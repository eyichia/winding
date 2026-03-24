"""江乙加Eric 最大排線程式""" 
import os
import sys
import io
import datetime
import openpyxl
import csv
import math
import json
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QMessageBox, QListView, QComboBox, QMenuBar, QMenu,
                                QDialog, QTableWidget, QTableWidgetItem, QVBoxLayout, QHeaderView, QFileDialog, QDoubleSpinBox, QSpinBox,
                                QLineEdit)
from PySide6.QtCore import Qt, QDateTime, QDate, QUrl, QRect, QPoint, QMarginsF, QBuffer, QByteArray
from PySide6.QtGui import (QIcon, QPixmap, QFont, QTextDocument, QPageLayout, QPageSize, QImage, QPen, 
                            QPdfWriter, QTextCursor, QTextTableFormat, QTextCharFormat, QTextImageFormat,
                            QColor, QTextBlockFormat, QTextTableCellFormat, QTextLength, QPainter)

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines # 如果需要精細控制格線
from openpyxl.chart.shapes import GraphicalProperties # 新增

from openpyxl.chart.layout import Layout, ManualLayout

from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (Paragraph, ParagraphProperties, CharacterProperties, 
                                   Font as DrawingFont, RegularTextRun)

from WireArrangeGUI_ui import Ui_Arrange
from Sub.utils import show_prompt_window




# """取得資源絕對路徑，兼容開發與 PyInstaller 打包模式"""
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller 運行時的臨時資料夾路徑
        return os.path.join(sys._MEIPASS, relative_path)
    # 開啟開發模式下的路徑
    return os.path.join(os.path.abspath("."), relative_path)
# """DATA 類別,長度"""
DATA_CONFIG = {
    "Data1": (1, 6),    # 起始, 數量
    "Data2": (101, 16)
}
# """主畫面"""
class MyMainWindow(QMainWindow, Ui_Arrange): 
    version = " v1.1"
    Developer = " 江乙加 Eric Chiang"
    ver_date = " 2026-03-19"
    Copyright = f" 2026 {Developer}" #" 2026 " + Developer
    
    
# """initial"""
    def __init__(self, parent=None):
        super(MyMainWindow, self).__init__(parent)
        self.setupUi(self)  # 這一行沒寫，提示絕對出不來
        # --- 【核心修改】預載入 Logo 資源 ---
        self.logo_pixmap = QPixmap(resource_path("Assets/Mylogo.png"))
        self.style_pixmap = QPixmap(resource_path("Assets/Mystyle.png"))
        self.current_lang = "TW" # 開機語言

        self.init_ui_settings() # 設定圖示與預設圖
        self.set_default_value() # 預設值
        self.load_languages_json() #載入語言json檔案
        self.update_global_reverse_map()
        #if not self.load_settings(): self.current_lang = "TW" # 設定初始語言，這要對應到你的檔名
        self.translate() #語言翻譯
        self.connect_signals() # 啟動所有按鈕與輸入框的「電線」連線！
        self.coords_calculation() # 座標計算

        self.Process.clear() # 清除紀錄
# """設定視窗圖示與初始排線圖"""
    def init_ui_settings(self):
        self.setWindowIcon(QIcon(self.style_pixmap)) #設定視窗圖示
        if hasattr(self, "Language"):
            self.Language.setIcon(QIcon(resource_path("Assets/Earth.png")))
            self.language_tc.setIcon(QIcon(resource_path("Assets/FlagTw.png")))
            self.language_sc.setIcon(QIcon(resource_path("Assets/FlagCn.png")))
            self.language_en.setIcon(QIcon(resource_path("Assets/FlagEn.png")))
        
        self.update_image("Arrange.png") # 預設顯示排線示意圖
# """預設值"""
    def set_default_value(self):
        self.in_pa_001.setValue(1.0000)
        self.in_pa_002.setValue(10.7951)
        self.in_pa_003.setValue(10.8723)
        self.in_pa_004.setValue(7.8818)
        self.in_pa_005.setValue(0.0000)
        self.in_pa_006.setValue(0.0000)
# """ 自動掃描 Languages 資料夾並載入所有 JSON """
    def load_languages_json(self):
        self.languages = {}
        lang_dir = resource_path("Languages") # 你的語言包資料夾
        if not os.path.exists(lang_dir):
            os.makedirs(lang_dir)
            return

        for filename in os.listdir(lang_dir):
            if filename.endswith(".json"):
                lang_key = filename.replace(".json", "") # 例如 "en_US"
                try:
                    with open(os.path.join(lang_dir, filename), 'r', encoding='utf-8') as f:
                        self.languages[lang_key] = json.load(f)
                except Exception as e:
                    print(f"Error loading {filename}: {e}")
# """ 從當前語言包抓取訊息文字 """
    def get_msg(self, key, default=""):
        lang = self.languages.get(self.current_lang, {})
        return lang.get("Messages", {}).get(key, default)
# """選單翻譯"""
    def translate_menu(self, menu_obj: QMenuBar | QMenu, lan_dict: dict):
        """ 
        純文字切換版：只負責更新 QAction/QMenu 的文字內容。
        """
        for action in menu_obj.actions():
            obj_name = action.objectName()
            submenu = action.menu()

            # 1. 處理 QAction 與 QMenu 的名稱連動偵測
            if not obj_name and submenu:
                obj_name = submenu.objectName()

            # 2. 匹配 JSON 並僅更新文字
            if obj_name in lan_dict:
                data = lan_dict[obj_name]
                # 相容兩種格式：如果是字典就拿 "text"，如果是純字串就直接用
                text = data["text"] if isinstance(data, dict) else data
                
                action.setText(text) # 更新 Action 文字
                if submenu:
                    submenu.setTitle(text) # 同步更新選單標題
                    # 遞迴翻譯子選單
                    self.translate_menu(submenu, lan_dict)
# """語言翻譯"""
    def translate(self):
        if not self.languages: return
        lang = self.languages.get(self.current_lang, {})
        # 1. 視窗標題
        self.setWindowTitle(lang.get("Window_main", {}).get("Winding", "繞線排線計算"))
        # 2. 選單列
        self.translate_menu(self.menuBar(), lang.get("Menubar", {}))
        # 3. 批次處理分類元件 (QLabel, QPushButton, QCombobox, Data1)
        t_default = lang.get("QLabel", {}).get("textFont", {"font_size": 14, "font_family": "Microsoft YaHei"})
        categories = ["QLabel", "QPushButton", "QCombobox"] + list(DATA_CONFIG.keys())
        for cat in categories:
            cat_data = lang.get(cat, {})
            for obj_name, data in cat_data.items():
                if obj_name == "textFont": continue # 跳過字體設定項
                
                widget = getattr(self, obj_name, None)
                if widget:
                    # 文字設定
                    content = data.get("text", "")
                    if isinstance(widget, QComboBox) and isinstance(content, list):
                        idx = widget.currentIndex()
                        widget.clear()
                        widget.addItems(content)
                        widget.setCurrentIndex(idx)
                    elif hasattr(widget, "setText"):
                        widget.setText(content)
                    
                    # 字體設定
                    f_size = data.get("font_size", t_default.get("font_size"))
                    f_family = data.get("font_family", t_default.get("font_family"))
                    widget.setFont(QFont(f_family, f_size))
# """繁簡英切換"""                
    def switch_language(self, new_lan):
        self.current_lang = new_lan
        self.translate()     
# """建立data1陣列，儲存所有「標籤、輸入物件、單位」的組合"""
    def data_list(self, category):
        lang = self.languages.get(self.current_lang, {})
        collected_data = {}

        if category not in DATA_CONFIG:
            print(f"錯誤：找不到分類 {category}")
            return {}
        start, count = DATA_CONFIG[category]

        for i in range(start, start + count):
            pa_label = f"in_pa_{i:03d}"
            # 1. 先抓到 Data1 裡面的那個小字典，如果找不到就給空字典 {}
            pa_dict = lang.get(category, {}).get(pa_label, {}) 
            # 2. 從小字典裡拿出 "text" 的內容，如果連小字典都沒有，就回傳 pa_key (如 lb_pa_001)
            pa_name = pa_dict.get("text", pa_label) if isinstance(pa_dict, dict) else pa_label 
            target_widget = getattr(self, pa_label, None)
            if target_widget and hasattr(target_widget, 'value'): # 檢查，這能確保萬一編號對應到的是 QLabel 或別的元件，程式不會崩潰。
                collected_data[pa_name] = target_widget.value()
        return collected_data    
# """集中處理所有按鈕與輸入框的連線"""
    def connect_signals(self):
        # --- 使用 lambda 精簡按鈕連線 ---
        # 格式：lambda: self.change_mode("圖片檔名", "模式名稱")
        #按鈕
        self.PB_chart.clicked.connect(lambda: self.change_mode("Arrange.png", "Diagram"))
        self.PB_chart_A.clicked.connect(lambda: self.change_mode("ArrangeA.png", "Type A"))
        self.PB_chart_B.clicked.connect(lambda: self.change_mode("ArrangeB.png", "Type B"))
        self.PB_chart_C.clicked.connect(lambda: self.change_mode("ArrangeC.png", "Type C"))
        self.PB_Export_coordinates.clicked.connect(self.export_coordinates)
        #self.PB_A_export_coordinates.clicked.connect(lambda: self.export_coordinates("A"))
        #self.PB_B_export_coordinates.clicked.connect(lambda: self.export_coordinates("B"))
        #self.PB_C_export_coordinates.clicked.connect(lambda: self.export_coordinates("C"))
        self.PB_Calculate.clicked.connect(self.arrange_calculation)
        self.PB_Clear_process.clicked.connect(self.Process_clear)
        #檔案選項
        self.file_save_csv.triggered.connect(self.data1_save_csv)
        self.file_open_csv.triggered.connect(self.data1_open_csv)
        self.file_save_txt.triggered.connect(self.data1_save_txt)
        self.file_open_txt.triggered.connect(self.data1_open_txt)
        self.file_export_csv.triggered.connect(self.export_summary_csv)
        self.file_export_pdf.triggered.connect(self.export_summary_pdf)
        self.file_export_xlsx.triggered.connect(self.export_summary_xlsx)
        self.file_exit.triggered.connect(self.close)
        self.file_reset.triggered.connect(self.set_default_value)

        self.language_tc.triggered.connect(lambda: self.switch_language("TW"))
        self.language_sc.triggered.connect(lambda: self.switch_language("CN"))
        self.language_en.triggered.connect(lambda: self.switch_language("EN"))
        self.about_version.triggered.connect(self.show_version)
        # --- 座標連動：改用 editingFinished (Enter 或 離開時觸發) ---
        spinbox_group = [
            self.in_pa_002, self.in_pa_003, self.in_pa_004, # AB, BC, CD 長度
            self.in_pa_005, self.in_pa_006                  # Bx, By 基準點
        ]
        for sb in spinbox_group:
            sb.editingFinished.connect(self.coords_calculation)    
# """通用切換模式功能"""
    def change_mode(self, img_file, mode_name):
        if not self.languages: return
        lang = self.languages.get(self.current_lang, {}).get("QPushButton", {})
        if mode_name == "Diagram":
            name = lang.get("PB_chart", {}).get("text", "排線示意")
        elif mode_name == "Type A":   
            name = lang.get("PB_chart_A", {}).get("text", "方式 A")
        elif mode_name == "Type B":   
            name = lang.get("PB_chart_B", {}).get("text", "方式 B")  
        else:
            name = lang.get("PB_chart_C", {}).get("text", "方式 C")     
        self.update_image(img_file)
        # 讀取並更新文字紀錄 (UI 中為 Process)
        now = datetime.datetime.now().strftime('%H:%M:%S')
        # 直接追加一行，不用先讀取舊文字再拼接
        self.Process.append(f"[{now}] {self.get_msg('change_mode', '切換至')} {name}")
# """通用換圖助手"""
    def update_image(self, filename):
        path = resource_path(f"Assets/{filename}")
        pixmap = QPixmap(path)
        if not pixmap.isNull():
            self.arrange_image.setPixmap(pixmap)
            self.arrange_image.setScaledContents(True)
        else:
            self.arrange_image.setText(f"找不到檔案: {filename}")
# """座標核心計算法"""
    def coords_calculation(self):
        try:
            # 1. 取得數值
            ab, bc, cd = self.in_pa_002.value(), self.in_pa_003.value(), self.in_pa_004.value()
            bx, by = self.in_pa_005.value(), self.in_pa_006.value()
            # 2. 定義座標映射表
            coord_map = {
                self.Coordinate_AX: bx,
                self.Coordinate_AY: ab + by,
                self.Coordinate_CX: bx + bc,
                self.Coordinate_CY: by,
                self.Coordinate_DX: bx + bc,
                self.Coordinate_DY: by + cd
            }
            # 3. 批量更新
            for widget, val in coord_map.items():
                widget.setValue(val)
            # 4. 紀錄
            self.Process.append(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {self.get_msg('coords_update', '座標已更新')}")
            # 5. 回傳座標資料
            return {
                "A": (bx, ab + by),
                "B": (bx, by),
                "C": (bx + bc, by),
                "D": (bx + bc, by + cd),
                "AB": ab,
                "BC": bc,
                "CD": cd
            }
        except Exception as e:
            print(f"計算錯誤: {e}")
# """計算所有組合"""
    def arrange_calculation(self):
        self.PB_Calculate.setEnabled(False) # 防止計算中又按一次按鈕,最後一行再Enable
        #1. 計算座標
        coords_data = self.coords_calculation()
        if not coords_data: return
        #2. 清除紀錄
        self.Process_clear()
        try:
            # 1. 取得數值
            wire_diam = self.in_pa_001.value() # 直徑
            wire_rad = wire_diam / 2 # 半徑
            cross_sectional_area = math.pi * (wire_rad**2) #線截面積            
            ab_l, bc_l, cd_l = coords_data["AB"], coords_data["BC"], coords_data["CD"]
            ax, ay = coords_data["A"] #coords_Ax, coords_Ay, coords_Dx, coords_Dy = coords_data["A"][0], coords_data["A"][1], coords_data["D"][0], coords_data["D"][1]
            bx, by = coords_data["B"]
            # 2. 排線幾何常數
            layer_pitch =(math.sqrt(3) / 2) * wire_diam #第2層後占用高度
            slope = (ay - coords_data["D"][1]) / (ax - coords_data["D"][0]) # AD斜率,斜率是負的，這條線在座標圖上是「從左上往右下」傾斜的
                       
            """計算相關數據**顯示用**"""
            usable_area = ((ab_l + cd_l) * bc_l) / 2 # 繞線梯形面積
            angle_ADa = math.degrees(math.atan2(ab_l - cd_l, bc_l)) # ∠ADa'
            total_layers_ab = 1 + int((ab_l - wire_diam) / (math.sqrt(3) * wire_rad)) #typeAB 總層數
            total_layers_c = int(ab_l / wire_diam) #typeC總層數
            square_layers_ab = 1 + int((cd_l - wire_diam) / (math.sqrt(3) * wire_rad)) #typeAB 方形層數
            square_layers_c = int(cd_l / wire_diam) #typeC 方形層數
            triangle_layers_ab = total_layers_ab - square_layers_ab #typeAB 三角形層數
            triangle_layers_c = total_layers_c - square_layers_c #typeC 三角形層數
            triangle_Height = ab_l - cd_l # 三角形高
            """----------"""
            # 3. 建立統一字典
            type_list = ["A", "B", "C"]
            _type = self.get_msg("report_footer1")
            _turns = self.get_msg("report_footer2")
            _rate = self.get_msg("report_footer3")
            self.winding_results = {}
            
            for type_name in type_list:
                turns = 0
                rate = 0.0
                self.winding_results[type_name] = {"total_turns": turns, "fill_rate": rate}
                square_layers = square_layers_c if type_name == "C" else square_layers_ab

                """建立type A方形各層資料{Type: {"層":(0:coordsX, 1:coordsY, 2:Turns, 3:Height)}}"""
                for i in range(1, square_layers + 1):
                    layer_key = f"L{i}"
                    if type_name == "A":
                        map_x = (2 * wire_rad) + bx if i % 2 == 0 else wire_rad + bx
                        map_y = wire_rad + by + (i - 1) * layer_pitch
                    elif type_name == "B":
                        map_x = wire_rad + bx if i % 2 == 0 else (2 * wire_rad) + bx
                        map_y = wire_rad + by + (i - 1) * layer_pitch
                    else:
                        map_x = wire_rad + bx   
                        map_y = wire_rad + by + (i - 1) * wire_diam
                    if type_name == "C":
                        map_turns = int(bc_l / wire_diam)
                        map_height = wire_diam + (i - 1) * wire_diam
                    else:
                        map_turns = int((bc_l - (map_x - wire_rad - bx)) / wire_diam)
                        map_height = wire_diam + (i - 1) * layer_pitch

                    # 寫入字典
                    self.winding_results[type_name][layer_key] = (map_x, map_y, map_turns) #(map_x, map_y, map_turns, map_height)
                    turns += map_turns
                    #self.winding_results[type_name]["total_turns"] = turns
                    # 顯示記錄值
                    self.Process.append(f"{_type}: {type_name}_ {layer_key}: ({map_x:.4f}, {map_y:.4f}, {map_turns:.0f})") #, {map_height:.4f})")

                square_last_height = cd_l - map_height#self.winding_results[type_name][f"L{square_layers}"][3] # 方形剩餘高度
                """建立type A三角形各層資料{"層":(0:coordsX, 1:coordsY, 2:Turns, 3:Height)}"""
                """
                AD直線方程式 : y - y1 = m(x - x1)
                >>> mx + y + k = 0
                >>> A = -m, B = 1, C = mx1 - y1  
                x1,y1用A(0, 10.7951)座標帶入 : y - 10.7951 = -0.2679(x - 0)
                >>> 0.2679 x + y - 10.7951= 0
                >>>A = 0.2679, B = 1, C = -10.7951
                圓心到CD距離d = |Axi + Byi + C| / √(A^2 + B^2)
                """
                eq_A, eq_B, eq_C = -slope, 1, (slope * ax -ay)
                triangle_layers = triangle_layers_c if type_name == "C" else triangle_layers_ab
                for j in range(1, triangle_layers + 1):
                    layer_num = square_layers + j # 三角形第一層=方形最總層數+1
                    layer_key = f"L{layer_num}"
                    if type_name == "A":
                        map_x = (2 * wire_rad) + bx if layer_num % 2 == 0 else wire_rad + bx
                        map_y = wire_rad + by + (layer_num - 1) * layer_pitch
                    elif type_name == "B":
                        map_x = wire_rad + bx if layer_num % 2 == 0 else (2 * wire_rad) + bx
                        map_y = wire_rad + by + (layer_num - 1) * layer_pitch
                    else:
                        map_x = wire_rad + bx  
                        map_y = wire_rad + by + (layer_num - 1) * wire_diam

                    # 利用 while 迴圈精確計算每一層能塞幾根，直到撞到 AD 斜邊
                    map_turns = 0
                    while True:
                        test_x = map_x + (map_turns * wire_diam)
                        # 點到直線距離公式
                        dist = abs(eq_A * test_x + eq_B * map_y + eq_C) / math.sqrt(eq_A**2 + eq_B**2)
                        # 判斷：不能超過右邊界 C 點，且不能撞到 AD 斜邊
                        if (test_x + wire_rad) > (bc_l + bx)  or dist < wire_rad:
                            break
                        map_turns += 1

                    if map_turns <= 0: break
                    map_height = (wire_diam - square_last_height) + (j - 1) * wire_diam if type_name == "C" else (layer_pitch - square_last_height) + (j - 1) * layer_pitch
                    self.winding_results[type_name][layer_key] = (map_x, map_y, map_turns)#, map_height)
                    turns += map_turns
                    self.Process.append(f"{_type}: {type_name}_ {layer_key}: ({map_x:.4f}, {map_y:.4f}, {map_turns:.0f})")#, {map_height:.4f})")
                self.winding_results[type_name]["total_turns"] = turns  
                rate = (turns * cross_sectional_area * 100) / usable_area # 槽滿率
                self.winding_results[type_name]["fill_rate"] = rate
                
                self.Process.append(f"{_type}: {type_name} {_turns}: {turns:.0f} {_rate}: {rate:.4f}")


            # 1. 找出總圈數最多的方式
            # 使用 max 函式搭配 lambda 來比較字典裡的數值
            best_type = max(self.winding_results, key=lambda t: self.winding_results[t]["total_turns"])
            best_turns = self.winding_results[best_type]["total_turns"]
            best_rate = self.winding_results[best_type]["fill_rate"]

            title = self.get_msg("end_calculation_title", "最優方案推薦")
            text = (
                f"<b>{self.get_msg('end_calculation_row1')}{best_type}</b><br>"
                f"<hr>" # 自動畫一條漂亮的水平線
                f"{self.get_msg('end_calculation_row2')}{best_turns:.0f}<br>"
                f"<br>"
                f"{self.get_msg('end_calculation_row3')}{best_rate:.4f}%<br>"
                f"<hr>"
                f"{self.get_msg('end_calculation_row4')}"
            )
            icon = self.logo_pixmap
            buttons = QMessageBox.StandardButton.Close
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)         
            """更新 UI 顯示"""
            self.Cross_sectional_area.setValue(cross_sectional_area) #線截面積
            self.Usable_area.setValue(usable_area) #繞線面積
        
        except Exception as e:
            print(f"計算錯誤: {e}")
        self.PB_Calculate.setEnabled(True)            
# """清除紀錄"""
    def Process_clear(self):
        title = self.get_msg("clear_title", "警告: 清除紀錄")
        text = self.get_msg("clear_text", "您確定要清除紀錄嗎？")
        icon = QMessageBox.Icon.Warning
        buttons = QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        theme = "alarm"
        font_size = 14
        icon_size = 80
        result = show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
        if result == QMessageBox.StandardButton.Yes or result == QMessageBox.StandardButton.Ok:
            self.Process.clear()
        #now = datetime.datetime.now().strftime('%H:%M:%S')
        #self.Process.append(f"[{now}] 紀錄已成功清空。")
# """data1 save csv"""
    def data1_save_csv(self):
        self.data1_params = self.data_list("Data1")
        _folder, _filename, _format, _title, _mode = "Data", "Winding", "csv", "csv_save_title", "save"
        file_path = self.data_processing(_folder, _filename, _format, _title, _mode)
        if not file_path: return  # <-- 預防使用者按取消

        #header = [self.get_msg("header_item", "項目"), self.get_msg("header_value", "數值")]
        data = []
        for label, val in self.data1_params.items():
            data.append([label, val])
        try:
            with open(file_path, mode='w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                #writer.writerow(header) # 務必寫入表頭
                writer.writerows(data)

            title = self.get_msg("csv_save_title", "儲存 CSV 數據")
            text = f"{self.get_msg('csv_save_ok', 'CSV 數據已存至：')}{file_path}"
            icon = self.logo_pixmap
            buttons = QMessageBox.StandardButton.Ok
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)   
        except Exception as e:
            title = self.get_msg("fail")
            text = f"{self.get_msg('儲存失敗：')}{str(e)}"
            icon = QMessageBox.Icon.Warning
            buttons = QMessageBox.StandardButton.Close
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """data1 open csv"""
    def data1_open_csv(self):
        #_folder, _filename, _format, _title, _mode = "Data", "Winding", "csv", "csv_save_title", "open"
        file_path = self.data_processing("Data", "Winding", "csv", "csv_open_title", "open")
        if not file_path: return  # <-- 預防使用者按取消

        #lang_data = self.languages.get(self.current_lang, {})
        #reverse_map = {}
        #for lang_code in self.languages:  # 遍歷所有已載入的語言
            #lang_data = self.languages[lang_code]
            #data1_dict = lang_data.get("Data1", {})
            #for i in range(1, 7):
                #pa_label = f"in_pa_{i:03d}"
                #translated_text = data1_dict.get(pa_label, {}).get("text")
                #if translated_text:
                    #reverse_map[translated_text] = pa_label
        # 3. 讀取 CSV 並填入數值
        try:
            with open(file_path, mode='r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                #next(reader)  # 跳過表頭 (項目, 數值)
                count = 0
                for row in reader:
                    if len(row) < 2: continue
                    #label = row[0] # CSV 裡的中文名稱
                    #value = row[1]    # CSV 裡的數值
                    
                    # 根據中文名稱找回對應的 widget 代碼
                    #widget_name = reverse_map.get(label)
                    widget_name = self.global_reverse_map.get(row[0])
                    if widget_name:
                        widget = getattr(self, widget_name, None)
                        if widget:
                            try:
                                widget.setValue(float(row[1]))
                                count += 1
                            except ValueError:
                                continue # 跳過無法轉換的行    
            self.coords_calculation()                    
            # 4. 成功提示
            title = self.get_msg("success", "成功")
            text = self.get_msg("updated", "成功讀取並更新")
            icon = self.logo_pixmap
            buttons = QMessageBox.StandardButton.Ok
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size) 

        except Exception as e:
            title = self.get_msg("fail")
            text = f"{self.get_msg('error')}{str(e)}"
            icon = QMessageBox.Icon.Warning
            buttons = QMessageBox.StandardButton.Close
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """data1 save txt"""
    def data1_save_txt(self):
        self.data1_params = self.data_list("Data1")
        _folder, _filename, _format, _title, _mode = "Data", "Winding", "txt", "txt_save_title", "save"
        file_path = self.data_processing(_folder, _filename, _format, _title, _mode)
        if not file_path: return

        data = []
        #data.append(f"{self.get_msg('header_item', '項目')}\t {self.get_msg('header_value', '數值')}")
        for label, val in self.data1_params.items():
            data.append(f"{label}\t: {val}") # 使用 \t (Tab) 取代逗號，數值會自動對齊到下一格縮排
        try:
            with open(file_path, mode='w', newline='', encoding='utf-8-sig') as f:
                writer = f.write
                writer("\n".join(data))

            self.coords_calculation() # 更新座標
            title = self.get_msg("txt_save_title", "儲存文字紀錄")
            text = f"{self.get_msg('txt_save_ok', 'TXT 數據已存至：')}{file_path}"
            icon = self.logo_pixmap
            buttons = QMessageBox.StandardButton.Ok
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)   
        except Exception as e:
            title = self.get_msg("fail")
            text = f"{self.get_msg('儲存失敗：')}{str(e)}"
            icon = QMessageBox.Icon.Warning
            buttons = QMessageBox.StandardButton.Close
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """data1 open txt"""
    def data1_open_txt(self):
        file_path = self.data_processing("Data", "Winding", "txt", "txt_open_title", "open")
        if not file_path: return  # <-- 預防使用者按取消
        # 2. 建立全域逆向索引 (同 CSV 邏輯)
        #global_reverse_map = {}
        #for lang_code in self.languages:
            #lang_data = self.languages[lang_code]
            #data1_dict = lang_data.get("Data1", {}) # json裡Data1
            #for i in range(1, 7):
                #pa_label = f"in_pa_{i:03d}"
                #translated_text = data1_dict.get(pa_label, {}).get("text")
                #if translated_text:
                    #global_reverse_map[translated_text] = pa_label

        # 3. 讀取 TXT
        try:
            count = 0
            with open(file_path, mode='r', encoding='utf-8-sig') as f:
                for line in f:
                    line = line.strip() # 去除結尾的換行符
                    if not line: continue
                    
                    # 判斷分隔符 (相容逗號、冒號或 Tab)
                    separator = ""
                    if ":" in line: separator = ":"
                    elif "," in line: separator = ","
                    elif "\t" in line: separator = "\t"
                    
                    if separator:
                        parts = line.split(separator)
                        if len(parts) >= 2:
                            label_txt = parts[0].strip()
                            value_txt = parts[1].strip()
                            
                            # 找回對應的 ID 並更新 UI
                            widget_name = self.global_reverse_map.get(label_txt)
                            if widget_name:
                                widget = getattr(self, widget_name, None)
                                if widget:
                                    widget.setValue(float(value_txt))
                                    count += 1
            
            self.coords_calculation() # 更新座標
            title = self.get_msg("success", "成功")
            text = self.get_msg("updated", "成功讀取並更新")
            icon = self.logo_pixmap
            buttons = QMessageBox.StandardButton.Ok
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size) 
                               
        except Exception as e:
            title = self.get_msg("fail")
            text = f"{self.get_msg('error')}{str(e)}"
            icon = QMessageBox.Icon.Warning
            buttons = QMessageBox.StandardButton.Close
            theme = "default"
            font_size = 12
            icon_size = 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """資料處理,路徑,資料夾,檔名,格式"""
    def data_processing(self, folder, filename, format, title, mode):
        # 1. 準備路徑與資料夾
        default_dir = os.path.join(os.getcwd(), folder)
        if not os.path.exists(default_dir): os.makedirs(default_dir)
        # 2. 準備過濾器
        filters = {
            "csv": "CSV Files (*.csv)",
            "xlsx": "xlsx Files (*.xlsx)",
            "txt": "Text Files (*.txt)",
            "pdf": "PDF Files (*.pdf)",
            "png": "Png Files (*.png)",
            "json": "JSON Files (*.json)"
        }
        file_filter = filters.get(format, "All Files (*)")
        # 3. 根據模式執行不同的對話框
        if mode == "save":
            # 儲存模式：自動生成帶時間戳的預設檔名
            default_filename = f"{filename}_{QDateTime.currentDateTime().toString('yyyyMMdd_HHmmss')}.{format}"
            initial_path = os.path.join(default_dir, default_filename)
            file_path, _ = QFileDialog.getSaveFileName(self, self.get_msg(title), initial_path, file_filter)
        else:
            # 開啟模式：直接打開資料夾，不需要預設檔名
            file_path, _ = QFileDialog.getOpenFileName(self, self.get_msg(title), default_dir, file_filter)
        # 4. 統一回傳檢查
        return file_path if file_path else None
# """ 建立全域逆向索引：將所有語言的標籤對應到元件 ID """
    def update_global_reverse_map(self):
        # 分類名稱 : (起始編號, 數量)
        data_config = {
            "Data1": (1, 6),   # 從 001 開始，抓 6 個
            "Data2": (101, 16) # 從 101 開始，抓 16 個
        }
        self.global_reverse_map = {}
        # 2. 開始遍歷所有語言
        for lang_code, lang_data in self.languages.items():  
            # 3. 第二層：遍歷每一個分類區塊 (Data1, Data2)
            for category, (start, count) in data_config.items():
                category_dict = lang_data.get(category, {}) # 取得如 lang["Data1"]
                # 4. 第三層：根據起始編號與數量生成對應 ID
                # range(1, 1 + 6) 會跑 1, 2, 3, 4, 5, 6
                for i in range(start, start + count):
                    pa_label = f"in_pa_{i:03d}" # 對應 UI 元件名稱 (如 in_pa_101)
                    pa_key = f"lb_pa_{i:03d}"   # 對應 JSON 內的標籤 Key (如 lb_pa_101)
                    
                    # 5. 抓取翻譯文字
                    # 注意：你的 JSON 結構是 {"lb_pa_xxx": {"text": "名稱", ...}}
                    item_info = category_dict.get(pa_label, {})
                    if isinstance(item_info, dict):
                        translated_text = item_info.get("text")
                        
                        if translated_text:
                            # 建立映射表：{"線徑": "in_pa_001", "Speed": "in_pa_101"}
                            self.global_reverse_map[translated_text] = pa_label

        #print(f"✅ 全域索引更新完成！共紀錄 {len(self.global_reverse_map)} 個跨語言標籤")
# """ 輸出橫向對比報表 (CSV 格式) """
    def export_summary_csv(self):
        # 1. 檢查是否有計算結果
        if not hasattr(self, 'winding_results') or not self.winding_results:
            title, text = self.get_msg("fail"), self.get_msg("report_fail", "請先進行計算！")
            icon, buttons = QMessageBox.Icon.Warning, QMessageBox.StandardButton.Close
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
            return
        # 2. 獲取存檔路徑 (利用你的萬用助手)
        file_path = self.data_processing("Reports", "Winding report", "csv", "report", "save")
        if not file_path: return
        # 3. 準備表頭 (從語系檔抓取)
        header = [
            self.get_msg("report_footer1", "方式"),
            self.get_msg("report_footer2", "總圈數"),
            self.get_msg("report_footer3", "槽滿率"),
            self.get_msg("report_footer4", "總層數")
        ]
        # 4. 整理數據列
        rows = []
        for type_name in ["A", "B", "C"]:
            res = self.winding_results.get(type_name, {})
            if not res: continue
            
            # 計算總層數：計算字典中以 'L' 開頭的 Key 數量
            total_layers = sum(1 for key in res.keys() if key.startswith('L'))
            # 組成一行數據
            rows.append([
                type_name,
                f"{res.get('total_turns', 0):.0f}",
                f"{res.get('fill_rate', 0):.4f}",
                str(total_layers)
            ])
        # 5. 寫入檔案
        try:
            with open(file_path, mode='w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(header) # 寫入標題
                writer.writerows(rows)  # 寫入所有方式的數據
            
            title, text = self.get_msg("success"), f"{self.get_msg('csv_save_ok', 'CSV 數據已存至：')}{file_path}"
            icon, buttons = self.logo_pixmap, QMessageBox.StandardButton.Ok
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
        except Exception as e:
            title, text = self.get_msg("fail"), f"{self.get_msg('儲存失敗：')}{str(e)}"
            icon, buttons = QMessageBox.Icon.Warning, QMessageBox.StandardButton.Close
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """ 輸出橫向對比報表 (pdf 格式) """
    def export_summary_pdf(self):
        # 1. 檢查是否有計算結果
        if not hasattr(self, 'winding_results') or not self.winding_results:
            title, text = self.get_msg("fail"), self.get_msg("report_fail", "請先進行計算！")
            icon, buttons = QMessageBox.Icon.Warning, QMessageBox.StandardButton.Close
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
            return
        # 2. 獲取存檔路徑 (利用你的萬用助手)
        file_path = self.data_processing("Reports", "Winding report", "pdf", "report", "save")
        if not file_path: return
        # 3. 準備 PDF 寫入器與文件
        printer = QPdfWriter(file_path)
        printer.setPageSize(QPageSize(QPageSize.A4))
        printer.setPageMargins(QMarginsF(5, 5, 5, 5)) # 稍微縮小邊距讓表格更寬
        # 強制設定解析度，讓 HTML 渲染更精準
        printer.setResolution(300)
        doc = QTextDocument()

        # 4. 準備 HTML 內容與表格數據
        title = self.get_msg("PDFReport", "排線報告")
        date_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")

        # 整理表格列
        table_rows = ""
        for type_name in ["A", "B", "C"]:
            res = self.winding_results.get(type_name, {})
            total_layers = sum(1 for key in res.keys() if key.startswith('L'))
            table_rows += f"""
                <tr>
                    <td>方式 {type_name}</td>
                    <td>{res.get('total_turns', 0):.0f}</td>
                    <td>{res.get('fill_rate', 0):.4f}%</td>
                    <td>{total_layers}</td>
                </tr>
            """

        # 5. 組合 HTML 樣式與結構
        html_content = f"""
        <html>
            <head>
                <style>
                    body {{ font-family: "Microsoft YaHei"; }}
                    /* 表格置中與樣式 */
                    table {{ 
                        border-collapse: collapse; 
                        width: 100%;
                        margin-top: 30pt;
                        margin-bottom: 30pt;
                    }}
                    th, td {{
                        border: 1px solid #000000; 
                        padding: 15pt; 
                        text-align: center; 
                        font-size: 18pt;
                    }}
                    th {{ background-color: #f2f2f2; font-weight: bold; font-size: 18pt; }}
                    .footer {{ text-align: right; font-size: 12pt; color: #555555; margin-top: 30pt; }}
                    .img-container {{ text-align: center; margin-top: 20pt; }}
                    .conclusion {{
                        width: 90%; 
                        margin: 20pt auto; 
                        padding: 15pt; 
                        border: 2pt solid #2c3e50; 
                        background-color: #f9f9f9; 
                        font-size: 18pt;
                    }}
                </style>
            </head>
            <body>
                <h1 style="font-size: 120pt; color: #000000; text-align: center; margin-top: 10pt; margin-bottom: 10pt;">
                    {title}
                </h1>
                <br>
                <div style="text-align: center; font-size: 14pt;">
                    <p><b>{self.get_msg('date', '日期')}:</b> {date_str}</p>
                    <p><b>{self.get_msg('dev_label', '開發者')}:</b> {self.Developer}</p>
                </div>    
                <hr>
                <table align="center" style="border-collapse: collapse; width: 90%; margin-top: 20pt; margin-bottom: 20pt;">
                    <tr>
                        <th>{self.get_msg('report_footer1', '方式')}</th>
                        <th>{self.get_msg('report_footer2', '總圈數')}</th>
                        <th>{self.get_msg('report_footer3', '槽滿率')}</th>
                        <th>{self.get_msg('report_footer4', '總層數')}</th>
                    </tr>
                    {table_rows}
                </table>
                <div class="img-container">
                    <h3>排線示意圖</h3>
                    <br>
                    <img src="diagram" width="300">
                    <br>
                    <hr>
                </div>
                <div class="footer">
                    <p>{self.get_msg('report_footer5', '本報告由排線工具自動生成')} | {self.version}</p>
                </div>
            </body>
        </html>
        """
        # 5. 嵌入圖片
        current_pixmap = self.arrange_image.pixmap()
        if current_pixmap:
            doc.addResource(QTextDocument.ImageResource, QUrl("diagram"), current_pixmap)

        # 6. 【關鍵修正】設定文檔寬度並執行列印
        doc.setHtml(html_content)
        # 設定文檔寬度與印表機的可列印寬度一致，避免內容縮在一角
        doc.setTextWidth(printer.width()) 
        doc.print_(printer) # 自動處理分頁與縮放

        # 成功提示
        title, text = self.get_msg("success"), f"{self.get_msg('pdf_save_ok', 'PDF 報告已存至：')}{file_path}"
        icon, buttons = self.logo_pixmap, QMessageBox.StandardButton.Ok
        theme, font_size, icon_size = "default", 12, 64
        show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """ 輸出專業 Excel 試算表報表 """
    def export_summary_xlsx(self):
        # 1. 檢查是否有計算結果
        if not hasattr(self, 'winding_results') or not self.winding_results:
            title, text = self.get_msg("fail"), self.get_msg("report_fail", "請先進行計算！")
            icon, buttons = QMessageBox.Icon.Warning, QMessageBox.StandardButton.Close
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
            return

        # 2. 獲取存檔路徑 (利用萬用助手，格式設為 xlsx)
        file_path = self.data_processing("Reports", "Winding report", "xlsx", "report", "save")
        if not file_path: return

        # 3. 建立活頁簿與工作表
        wb = Workbook()
        ws = wb.active
        ws.title = self.get_msg("xlsx_title", "排線計算摘要")

        # 4. 定義樣式 (加粗、置中、框線)
        header_font = Font(name='Microsoft YaHei', size=14, bold=True, color="FFFFFF")
        cell_font = Font(name='Microsoft YaHei', size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        thin_side = Side(border_style="thin", color="000000")
        full_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

        # 5. 寫入標題與表頭
        ws.merge_cells('A1:D1') # 合併儲存格
        ws['A1'] = self.get_msg("PDFReport", "排線報告")
        ws['A1'].font = Font(name='Microsoft YaHei', size=20, bold=True)
        ws['A1'].alignment = center_align

        headers = [
            self.get_msg("report_footer1", "方式"),
            self.get_msg("report_footer2", "總圈數"),
            self.get_msg("report_footer3", "槽滿率"),
            self.get_msg("report_footer4", "總層數")
        ]
        # col 行[直A,B,C...], row 列[橫1,2,3...]
        for col, text in enumerate(headers, 1): # 處理A3~D3樣式
            cell = ws.cell(row=3, column=col, value=text)
            cell.font = header_font         # 設定字體（如：微軟正黑體、加粗、白色）
            cell.alignment = center_align   # 設定對齊（水平置中、垂直置中）
            cell.fill = header_fill         # 設定填充背景色（如：深藍色）
            cell.border = full_border       # 設定框線（四邊加上細實線）
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20 # 自動調整欄位寬度

        # 6. 寫入計算數據
        current_row = 4
        for type_name in ["A", "B", "C"]:
            res = self.winding_results.get(type_name, {})
            total_layers = sum(1 for key in res.keys() if key.startswith('L'))
            
            data = [
                f"方式 {type_name}",
                res.get('total_turns', 0),
                f"{res.get('fill_rate', 0):.4f}%",
                total_layers
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = full_border
            current_row += 1

        # 7. 插入排線示意圖 (關鍵步驟)
        #current_pixmap = self.arrange_image.pixmap()
        current_pixmap = QPixmap(resource_path(f"Assets/ArrangeA.png"))
        if current_pixmap:
            # 必須確保有 import openpyxl.drawing.image
            ba = QByteArray()
            buffer = QBuffer(ba)
            buffer.open(QBuffer.WriteOnly)
            
            # 將圖片存入 QBuffer
            current_pixmap.toImage().save(buffer, "PNG")
            
            # 交給 io.BytesIO，再由 ExcelImage (此時會呼叫 Pillow) 讀取
            img_data = io.BytesIO(ba.data())
            img = ExcelImage(img_data) # <--- 這裡會觸發對 Pillow 的呼叫
            
            img.width, img.height = 400, 400
            ws.add_image(img, 'F3')

        # 8. 儲存檔案
        try:
            wb.save(file_path)
            title, text = self.get_msg("success"), f"{self.get_msg('xlsx_save_ok', 'xlsx 數據已存至：')}{file_path}"
            icon, buttons = self.logo_pixmap, QMessageBox.StandardButton.Ok
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size) 
        except Exception as e:
            title, text = self.get_msg("fail"), f"{self.get_msg('儲存失敗：')}{str(e)}"
            icon, buttons = QMessageBox.Icon.Warning, QMessageBox.StandardButton.Close
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """ 輸出座標 Excel 試算表報表 """
    def export_coordinates(self):
        # 1. 檢查是否有計算結果
        if not hasattr(self, 'winding_results') or not self.winding_results:
            title, text = self.get_msg("fail"), self.get_msg("report_fail", "請先進行計算！")
            icon, buttons = QMessageBox.Icon.Warning, QMessageBox.StandardButton.Close
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
            return
        # 2. 獲取存檔路徑 (利用你的萬用助手)
        file_path = self.data_processing("Reports", "Winding coordinates", "xlsx", "report", "save")
        if not file_path: return

        # 3. 直接建立新活頁簿 (取消模板判斷)
        sheet_title = ["A", "B", "C"]
        sheet_name = self.get_msg("coordinates_title", "座標摘要")
        wb = Workbook()
        
        
          

        def set_sheet_contents(type):
            # 4. 定義樣式 (加粗、置中、框線)
            header_font = Font(name='Microsoft YaHei', size=14, bold=True, color="FFFFFF")
            cell_font = Font(name='Microsoft YaHei', size=12)
            center_align = Alignment(horizontal='center', vertical='center')
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            thin_side = Side(border_style="thin", color="000000")
            full_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
            # 5. 寫入標題與表頭
            ws.merge_cells('A1:B1') # 合併儲存格
            ws['A1'] = f"{self.get_msg('report_footer1')} {type}"
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = center_align
            ws.merge_cells('A9:C9') # 合併儲存格
            ws['A9'] = self.get_msg("coordinates", "座標(X,Y)")
            ws['A9'].font = header_font
            ws['A9'].fill = header_fill
            ws['A9'].alignment = center_align

            ws['S1'] = f"{self.get_msg('diagram')}"
            ws['S1'].font = cell_font
            ws['S1'].alignment = center_align

            if not self.languages: return
            lang = self.languages.get(self.current_lang, {}).get("Data1", {})
            headers = [
                self.get_msg("report_footer2", "總圈數"),
                self.get_msg("report_footer3", "槽滿率"),
                self.get_msg("report_footer4", "總層數"),
                lang.get("lb_pa_001", {}).get("text", "線徑"),
                lang.get("lb_pa_002", {}).get("text", "AB長度"),
                lang.get("lb_pa_003", {}).get("text", "BC長度"),
                lang.get("lb_pa_004", {}).get("text", "CD長度")
            ]
            for i in range(1, 4):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20 # 自動調整欄ABC位寬度
            # col 行[直A,B,C...], row 列[橫1,2,3...]
            # ws.cell(row=row, column=2, value=text) 
            # 寫入「列 (Row)」：定住 row，讓 column 隨迴圈增加。
            # 寫入「行 (Column)」：定住 column，讓 row 隨迴圈增加。
            current_row = 2
            for row, text in enumerate(headers, 1): # 處理A2~A8樣式 1>>從A欄
                cell = ws.cell(row=current_row, column=1, value=text)
                cell.font = cell_font         # 設定字體（如：微軟正黑體、加粗、白色）
                cell.alignment = center_align   # 設定對齊（水平置中、垂直置中）
                #cell.fill = header_fill         # 設定填充背景色（如：深藍色）
                cell.border = full_border       # 設定框線（四邊加上細實線）
                current_row += 1    
        
            # 6. 寫入計算數據
            res = self.winding_results.get(type, {})
            total_layers = sum(1 for key in res.keys() if key.startswith('L'))
            data = [
                res.get("total_turns", 0),
                f"{res.get('fill_rate', 0):.4f}%",
                total_layers,
                self.in_pa_001.value(),
                self.in_pa_002.value(),
                self.in_pa_003.value(),
                self.in_pa_004.value()
            ]
            current_row = 2
            for row, value in enumerate(data, 2): # 處理B2~B8樣式 2>>從B欄
                cell = ws.cell(row=current_row, column=2, value=value)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = full_border
                current_row += 1   

            # 7. 整理並寫入座標數據
            coords_item = [[],[],[]]    
            wire_diam = self.in_pa_001.value() # 直徑
            row_count = 0
            for i in range(1, total_layers+1):
                layer_key = f"L{i}"
                coords_x, coords_y, layer_turns = res[layer_key] # 取得座標X,Y與該總層圈數
                if i % 2 == 0: coords_x = coords_x + (layer_turns - 1 ) * wire_diam # 奇偶數層 

                if layer_key in res: # 確保 res 裡有這層資料再繼續
                    offset_x = 0
                    for j in range(1, int(layer_turns)+1):
                        coords_item[0].append(f"{layer_key}-{j}")
                        if not i % 2 == 0: # 奇數層 
                            coords_item[1].append(round(coords_x + offset_x, 4))
                        else: # 偶數層 
                            coords_item[1].append(round(coords_x - offset_x, 4))
                        coords_item[2].append(round(coords_y, 4))
                        offset_x += wire_diam
                        row_count += 1

            start_row = 10
            end_row = start_row + row_count - 1
            for col_idx, sub_list in enumerate(coords_item, 1): # 處理A10~An樣式 1>>從A欄  
                for row_offset, value in enumerate(sub_list):
                    # 計算實際行號：10 + 0, 10 + 1, 10 + 2...
                    target_row = start_row + row_offset  
                    # 寫入儲存格
                    cell = ws.cell(row=target_row, column=col_idx, value=value)
                    cell.font = cell_font         # 設定字體（如：微軟正黑體、加粗、白色）
                    cell.alignment = center_align   # 設定對齊（水平置中、垂直置中）
        

            # --- 8. 建立圖表 (樣式完全模擬模板) ---
            x_values = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
            y_values = Reference(ws, min_col=3, min_row=start_row, max_row=end_row)

            chart = ScatterChart()
            chart.scatterStyle = 'lineMarker'
            chart.style = 2
            chart.legend = None
            chart.height = 18 # 18cm 在drawing1.xml <xdr:ext cx="7920000" cy="6480000"/>
            chart.width = 22
            # 先初始化標題，使其不再是 None】
            chart.title = " " 
            chart.x_axis.title = " "
            chart.y_axis.title = " "

            # 文字格式設定助手 (16pt, 微軟正黑體)
            # # 注意寫在def export_coordinates裡面
            def set_16pt_style(target, text):
                font_style = DrawingFont(typeface='Microsoft YaHei')
                cp = CharacterProperties(sz=1600, b=False, latin=font_style, ea=font_style)
                pp = ParagraphProperties(defRPr=cp)
                run = RegularTextRun(t=text)
                rtp = Paragraph(pPr=pp, r=[run])
                target.tx = Text(rich=RichText(p=[rtp]))

            # 設定標題與軸名稱 (16pt)
            set_16pt_style(chart.title, self.get_msg("chart_title", "排線路徑座標圖"))
            set_16pt_style(chart.x_axis.title, self.get_msg("axis_x", "X 座標 (mm)"))
            set_16pt_style(chart.y_axis.title, self.get_msg("axis_y", "Y 座標 (mm)"))

            # 佈局與邊距設定
            chart.layout = Layout(
                manualLayout=ManualLayout(
                    x=0.15, y=0.10, h=0.75, w=0.85,
                    xMode="edge", yMode="edge", layoutTarget="inner"
                )
            )    

            # 刻度標籤字體設定 (10pt)
            tick_props = CharacterProperties(sz=1000)
            tick_para = ParagraphProperties(defRPr=tick_props)
            chart.x_axis.txPr = RichText(p=[Paragraph(pPr=tick_para, endParaRPr=tick_props)])
            chart.y_axis.txPr = RichText(p=[Paragraph(pPr=tick_para, endParaRPr=tick_props)])
            chart.x_axis.delete = False
            chart.y_axis.delete = False

            # 座標軸刻度邏輯
            chart.x_axis.majorUnit = 0.5        # 主要刻度間隔: 0.5
            chart.x_axis.minorUnit = 0.1        # 次要刻度間隔: 0.1
            chart.x_axis.majorTickMark = 'cross' # 主要刻度線: 交叉
            chart.x_axis.minorTickMark = 'in'    # 次要刻度線: 內側
            chart.x_axis.tickLblPos = 'nextTo'   # 標籤位置: 軸旁
            chart.x_axis.crosses = 'autoZero'    # 座標交叉點: 自動歸零
            chart.x_axis.crossBetween = 'midCat'
            chart.x_axis.number_format = '#,##0.0_);[Red]\(#,##0.0\)'
            chart.x_axis.majorGridlines = None
        
            chart.y_axis.majorUnit = 0.5        # 主要刻度間隔: 0.5
            chart.y_axis.minorUnit = 0.1        # 次要刻度間隔: 0.1
            chart.y_axis.majorTickMark = 'cross' # 主要刻度線: 交叉
            chart.y_axis.minorTickMark = 'in'    # 次要刻度線: 內側
            chart.y_axis.tickLblPos = 'nextTo'   # 標籤位置: 軸旁
            chart.y_axis.crosses = 'autoZero'    # 座標交叉點: 自動歸零
            chart.y_axis.number_format = '0.0'   # Y 軸數字格式：固定一位小數
            # 格線設定：僅保留 Y 軸水平線 (0.5pt)
            chart.y_axis.majorGridlines = ChartLines()
            chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(w=6350))
            chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(w=6350))
        
            # 數據序列樣式
            series = Series(y_values, x_values)
            series.marker = Marker(symbol='circle', size=5)
            line_color = "2E75B6"
            series.smooth = False
            series.graphicalProperties.line.solidFill = line_color
            series.graphicalProperties.line.width = 19050 # 1.5pt
            series.marker.graphicalProperties.solidFill = line_color
            series.marker.graphicalProperties.line.solidFill = line_color

            chart.series.append(series)
            chart.varyColors = False
            ws.add_chart(chart, "E2")

            # 插入排線示意圖
            current_pixmap = QPixmap(resource_path(f"Assets/Arrange{type}.png"))
            if current_pixmap:
                # 必須確保有 import openpyxl.drawing.image
                ba = QByteArray()
                buffer = QBuffer(ba)
                buffer.open(QBuffer.WriteOnly)
                # 將圖片存入 QBuffer
                current_pixmap.toImage().save(buffer, "PNG")
                # 交給 io.BytesIO，再由 ExcelImage (此時會呼叫 Pillow) 讀取
                img_data = io.BytesIO(ba.data())
                img = ExcelImage(img_data) # <--- 這裡會觸發對 Pillow 的呼叫
                img.width, img.height = 600, 600
                ws.add_image(img, 'S2')

        for idx, title in enumerate(sheet_title):
            if idx == 0: 
                ws = wb.active
                ws.title = f"{sheet_name}{title}"
                set_sheet_contents(title)
            else:
                ws = wb.active
                ws = wb.create_sheet(title=f"{sheet_name}{title}", index=idx)
                set_sheet_contents(title)
            print(ws)          

        
        
       
        # 8. 儲存檔案
        try:
            wb.save(file_path)
            title, text = self.get_msg("success"), f"{self.get_msg('xlsx_save_ok', 'xlsx 數據已存至：')}{file_path}"
            icon, buttons = self.logo_pixmap, QMessageBox.StandardButton.Ok
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size) 
        except Exception as e:
            title, text = self.get_msg("fail"), f"{self.get_msg('儲存失敗：')}{str(e)}"
            icon, buttons = QMessageBox.Icon.Warning, QMessageBox.StandardButton.Close
            theme, font_size, icon_size = "default", 12, 64
            show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)     





# """顯示軟體資訊視窗，且只改變此視窗的字型"""
    def show_version(self):
        title = self.get_msg("about_title", "關於排線計算程式")
        # 使用 HTML 語法設定內容
        text = (
            f"<h3>{self.get_msg('soft_name')}{self.version}</h3>"
            f"<p>{self.get_msg('dev_label')}{self.Developer}</p>"
            f"<p>{self.get_msg('date_label')}{self.ver_date}</p>"
            "<hr>"
            f"<p>{self.get_msg('about_desc')}</p>"
            f"<p><i>{self.get_msg('copyright')}{self.Copyright}</i></p>"
        )
        icon = self.logo_pixmap
        buttons = QMessageBox.StandardButton.Close
        theme = "default"
        font_size = 14
        icon_size = 64
        show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
# """關閉程式"""    
    def closeEvent(self, event):
        title = self.get_msg("close_title", "警告: 關閉程式")
        text = self.get_msg("close_text", "您確定要關閉程式嗎？")
        icon = QMessageBox.Icon.Question
        buttons = QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        theme = "alarm"
        font_size = 14
        icon_size = 80
        result = show_prompt_window(self, title, text, icon, buttons, theme, font_size, icon_size)
        if result == QMessageBox.StandardButton.Yes or result == QMessageBox.StandardButton.Ok:
            event.accept()  # 接受關閉事件，程式結束
        else:
            event.ignore()  # 忽略關閉事件，回到程式


if __name__ == "__main__":
    app = QApplication(sys.argv)
    try:
        myWin = MyMainWindow()
        myWin.show()
        sys.exit(app.exec())
    except Exception as e:
        print(f"啟動失敗: {e}")